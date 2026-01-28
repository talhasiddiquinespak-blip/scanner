import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

REGISTRY_DIR = "backend/registry"
REGISTRY_FILE = os.path.join(REGISTRY_DIR, "letter_registry.xlsx")


def ensure_registry_exists():
    """Create Excel registry if it doesn't exist."""
    if not os.path.exists(REGISTRY_DIR):
        os.makedirs(REGISTRY_DIR)

    if not os.path.exists(REGISTRY_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Letters"

        ws.append([
            "ID",
            "Date",
            "From",
            "To",
            "Subject",
            "Saved File",
            "Timestamp"
        ])

        wb.save(REGISTRY_FILE)


def append_letter_record(data: dict, file_name: str):
    """
    Append one scanned letter to Excel registry.
    """
    ensure_registry_exists()

    wb = load_workbook(REGISTRY_FILE)
    ws = wb.active

    next_id = ws.max_row
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.append([
        next_id,
        data.get("date", ""),
        data.get("from", ""),
        data.get("to", ""),
        data.get("subject", ""),
        file_name,
        timestamp
    ])

    wb.save(REGISTRY_FILE)
